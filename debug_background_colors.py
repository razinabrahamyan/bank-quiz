import openpyxl
import os

def debug_background_colors():
    """Debug background colors in 6.4.xlsx to understand correct answer detection"""
    
    file_path = 'public/xlsx/6.4.xlsx'
    
    if not os.path.exists(file_path):
        print(f"File {file_path} not found!")
        return
    
    print(f"Debugging background colors: {file_path}")
    print("=" * 80)
    
    try:
        # Load workbook with formatting
        wb = openpyxl.load_workbook(file_path, data_only=False)
        ws = wb.active
        
        print(f"Sheet name: {ws.title}")
        print("=" * 80)
        
        # Look for cells with content in column E and check their background colors
        answer_cells = []
        
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=5)  # Column E
            if cell.value and isinstance(cell.value, str) and cell.value.strip():
                bg_color = None
                if cell.fill and cell.fill.bgColor:
                    bg_color = cell.fill.bgColor.indexed
                
                answer_cells.append({
                    'row': row,
                    'value': cell.value.strip(),
                    'bg_color': bg_color
                })
        
        print(f"Found {len(answer_cells)} cells with content in column E")
        print("=" * 80)
        
        # Group by questions (look for patterns)
        current_question = None
        question_answers = []
        
        for i, cell_info in enumerate(answer_cells):
            value = cell_info['value']
            
            # Check if this looks like a question (starts with "Նշել" or "Համաձայն")
            if value.startswith('Նշել') or value.startswith('Համաձայն'):
                if current_question and question_answers:
                    print(f"\nQuestion: {current_question}")
                    for j, answer in enumerate(question_answers):
                        print(f"  Answer {j+1}: {answer['value']} (bg_color: {answer['bg_color']})")
                    print()
                
                current_question = value
                question_answers = []
            else:
                # This is an answer
                question_answers.append(cell_info)
        
        # Print the last question
        if current_question and question_answers:
            print(f"\nQuestion: {current_question}")
            for j, answer in enumerate(question_answers):
                print(f"  Answer {j+1}: {answer['value']} (bg_color: {answer['bg_color']})")
        
    except Exception as e:
        print(f"Error processing file: {e}")

if __name__ == "__main__":
    debug_background_colors() 