import pandas as pd
import openpyxl
from openpyxl import load_workbook
import json
import os

def debug_6_4():
    """Debug the 6.4.xlsx file to see why only 15 questions are found instead of 17"""
    
    file_path = 'public/xlsx/6.4.xlsx'
    
    if not os.path.exists(file_path):
        print(f"File {file_path} not found!")
        return
    
    print(f"Debugging: {file_path}")
    print("=" * 80)
    
    try:
        # Load workbook with data_only=True to get actual values
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb.active
        
        print(f"Sheet name: {ws.title}")
        print(f"Sheet dimensions: {ws.dimensions}")
        print("=" * 80)
        
        # Get all data
        data = []
        for row in ws.iter_rows(values_only=True):
            data.append(row)
        
        print(f"Total rows: {len(data)}")
        print("=" * 80)
        
        # Debug the state machine logic
        step = "group_title"  # Initial step
        current_group_title = None
        current_question = None
        current_answers = []
        question_count = 0
        
        print("DEBUGGING STATE MACHINE:")
        print("=" * 80)
        
        # Loop through all rows starting from index 1 (second row)
        for i in range(1, len(data)):
            row = data[i]
            row_num = i + 1  # 1-indexed row number
            
            print(f"\nRow {row_num}: {row}")
            
            if step == "group_title":
                # Look for question group title in Column B (index 1)
                if row and len(row) > 1 and row[1] and isinstance(row[1], str) and row[1].strip():
                    current_group_title = str(row[1]).strip()
                    print(f"  → Found question group: {current_group_title}")
                    print(f"  → Changing step from 'group_title' to 'question'")
                    step = "question"
                else:
                    print(f"  → No group title found in Column B (value: {row[1] if row and len(row) > 1 else 'None'})")
            
            elif step == "question":
                # Look for question in Column E (index 4)
                if row and len(row) > 4 and row[4] and isinstance(row[4], str) and row[4].strip():
                    current_question = str(row[4]).strip()
                    print(f"  → Found question: {current_question[:100]}...")
                    print(f"  → Changing step from 'question' to 'answers'")
                    step = "answers"
                    current_answers = []
                else:
                    print(f"  → No question found in Column E (value: {row[4] if row and len(row) > 4 else 'None'})")
            
            elif step == "answers":
                # Look for answers in Column E (index 4)
                if row and len(row) > 4 and row[4] and isinstance(row[4], str) and row[4].strip():
                    answer_text = str(row[4]).strip()
                    current_answers.append(answer_text)
                    print(f"  → Found answer {len(current_answers)}: {answer_text[:50]}...")
                    
                    # If we found 4 answers, save the question and reset
                    if len(current_answers) >= 4:
                        question_count += 1
                        print(f"  → COMPLETED QUESTION #{question_count}")
                        print(f"  → Group: {current_group_title}")
                        print(f"  → Question: {current_question[:100]}...")
                        print(f"  → Answers: {len(current_answers)} answers")
                        print(f"  → Changing step from 'answers' to 'group_title'")
                        
                        # Reset for next question
                        step = "group_title"
                        current_question = None
                        current_answers = []
                    else:
                        print(f"  → Still collecting answers ({len(current_answers)}/4)")
                else:
                    print(f"  → No answer found in Column E (value: {row[4] if row and len(row) > 4 else 'None'})")
        
        print("\n" + "=" * 80)
        print(f"DEBUG SUMMARY:")
        print(f"Total questions found: {question_count}")
        print(f"Final step: {step}")
        print(f"Current group title: {current_group_title}")
        print(f"Current question: {current_question}")
        print(f"Current answers count: {len(current_answers)}")
        
    except Exception as e:
        print(f"Error processing file: {e}")

if __name__ == "__main__":
    debug_6_4() 