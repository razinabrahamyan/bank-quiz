import pandas as pd
import openpyxl
from openpyxl import load_workbook
import json
import os

def process_excel_files():
    """Process Excel files and extract questions with answers"""
    questions = []
    
    # Look for Excel files in the public/xlsx directory
    excel_dir = 'public/xlsx'
    if not os.path.exists(excel_dir):
        print(f"Directory {excel_dir} not found. Creating it...")
        os.makedirs(excel_dir, exist_ok=True)
        print(f"Created directory {excel_dir}")
        print("Please place your Excel files in this directory and run the script again.")
        return questions
    
    excel_files = []
    for file in os.listdir(excel_dir):
        if file.endswith(('.xlsx', '.xls')):
            excel_files.append(os.path.join(excel_dir, file))
    
    if not excel_files:
        print(f"No Excel files found in {excel_dir}")
        print("Please place your Excel files (.xlsx or .xls) in the public/xlsx directory.")
        return questions
    
    print(f"Found {len(excel_files)} Excel files: {[os.path.basename(f) for f in excel_files]}")
    
    processed_files = 0
    questions_per_file = {}
    
    for file in excel_files:
        print(f"Processing {os.path.basename(file)}...")
        file_questions = 0
        
        try:
            # Load workbook with formatting to access cell properties
            wb = openpyxl.load_workbook(file, data_only=False)
            ws = wb.active
            
            # Get all data with cell objects to access formatting
            data = []
            for row in ws.iter_rows():
                data.append(row)
            
            print(f"  Loaded {len(data)} rows from {os.path.basename(file)}")
            
            # Skip first row (header) and start from row 2
            step = "group_title"  # Initial step
            current_group_title = None
            current_question = None
            current_answers = []
            current_answer_cells = []  # Track cells to check background color
            
            # Loop through all rows starting from index 1 (second row)
            for i in range(1, len(data)):
                row = data[i]
                
                if step == "group_title":
                    # Look for question group title in Column B (index 1)
                    if row and len(row) > 1 and row[1] and isinstance(row[1].value, str) and row[1].value.strip():
                        current_group_title = str(row[1].value).strip()
                        print(f"    Found question group: {current_group_title}")
                        
                        # Check if this same row also has a question in Column E
                        if row and len(row) > 4 and row[4] and isinstance(row[4].value, str) and row[4].value.strip():
                            current_question = str(row[4].value).strip()
                            print(f"    Found question in same row: {current_question[:100]}...")
                            step = "answers"
                            current_answers = []
                            current_answer_cells = []
                        else:
                            step = "question"
                
                elif step == "question":
                    # Look for question in Column E (index 4)
                    if row and len(row) > 4 and row[4] and isinstance(row[4].value, str) and row[4].value.strip():
                        current_question = str(row[4].value).strip()
                        print(f"    Found question: {current_question[:100]}...")
                        step = "answers"
                        current_answers = []
                        current_answer_cells = []
                
                elif step == "answers":
                    # Look for answers in Column E (index 4)
                    if row and len(row) > 4 and row[4] and isinstance(row[4].value, str) and row[4].value.strip():
                        answer_text = str(row[4].value).strip()
                        current_answers.append(answer_text)
                        current_answer_cells.append(row[4])  # Store cell for background check
                        print(f"    Found answer {len(current_answers)}: {answer_text[:50]}...")
                        
                        # If we found 4 answers, save the question and reset
                        if len(current_answers) >= 4:
                            # Determine correct answer by checking background color
                            correct_answer_index = 0  # Default to first answer
                            for j, cell in enumerate(current_answer_cells):
                                try:
                                    if (cell.fill and cell.fill.bgColor and 
                                        hasattr(cell.fill.bgColor, 'indexed') and 
                                        cell.fill.bgColor.indexed == 64):
                                        correct_answer_index = j
                                        break
                                except:
                                    # If there's an error checking background color, continue
                                    continue
                            
                            question_obj = {
                                "question": current_question,
                                "answers": current_answers,
                                "correctAnswerIndex": correct_answer_index,
                                "questionGroup": current_group_title,
                                "sourceFile": os.path.basename(file)
                            }
                            
                            questions.append(question_obj)
                            file_questions += 1
                            
                            print(f"    Extracted question with {len(current_answers)} answers, correct answer: {correct_answer_index}")
                            
                            # Reset for next question
                            step = "group_title"
                            current_question = None
                            current_answers = []
                            current_answer_cells = []
            
            questions_per_file[os.path.basename(file)] = file_questions
            processed_files += 1
            
        except Exception as e:
            print(f"Error processing {file}: {e}")
    
    # Save to JSON file
    output_file = 'public/questions.json'
    try:
        with open(output_file, 'w', encoding='utf-8') as f:
            json.dump(questions, f, ensure_ascii=False, indent=2)
        
        print(f"\nProcessing complete!")
        print(f"Total questions extracted: {len(questions)}")
        print(f"Files processed: {processed_files}")
        print(f"Questions per file:")
        for file, count in questions_per_file.items():
            print(f"  {file}: {count} questions")
        print(f"Output saved to: {output_file}")
        
    except Exception as e:
        print(f"Error saving to {output_file}: {e}")
    
    return questions

def main():
    """Main function to process Excel files"""
    questions = process_excel_files()
    return questions

if __name__ == "__main__":
    main() 